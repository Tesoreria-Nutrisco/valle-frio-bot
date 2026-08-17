import logging
import re
from pathlib import Path
from datetime import datetime
import pdfplumber
from pypdf import PdfReader

logger = logging.getLogger(__name__)


def extraer_metadatos_nomina(pdf_path):
    """
    Extrae metadatos de la nómina usando extracción de TABLAS estructuradas.

    Busca la tabla que contiene "ID nómina" y extrae:
    - "ID nómina"
    - "Fecha carga"
    - "Fecha pago"
    - "Estado"

    Returns:
        {
            'id_nomina': str,
            'fecha_carga': date,
            'fecha_pago': date,
            'estado': str
        }
    """
    pdf_path = Path(pdf_path)
    logger.info(f"Extrayendo metadatos de nómina (tablas): {pdf_path}")

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            if len(pdf.pages) == 0:
                raise ValueError("PDF vacío")

            page = pdf.pages[0]
            tables = page.extract_tables()

            if not tables:
                raise ValueError("No se encontraron tablas en la primera página")

            logger.info(f"Se encontraron {len(tables)} tabla(s) en página 1")

            id_nomina = None
            fecha_carga = None
            fecha_pago = None
            estado = None

            for table_idx, table in enumerate(tables):
                logger.debug(f"Tabla {table_idx}: {table}")

                for row_idx, row in enumerate(table):
                    row_str = str(row).lower()

                    if "id nómina" in row_str:
                        logger.info(f"Tabla {table_idx}, fila {row_idx}: encontrada 'ID nómina'")

                        header_row = None
                        data_row = None

                        if row_idx == 0 and "id nómina" in str(row[0]).lower():
                            header_row = row
                            if row_idx + 1 < len(table):
                                data_row = table[row_idx + 1]
                        elif row_idx > 0:
                            header_row = table[row_idx - 1] if row_idx > 0 else None
                            data_row = row

                        if header_row and data_row:
                            logger.debug(f"Header: {header_row}")
                            logger.debug(f"Data: {data_row}")

                            try:
                                idx_id = next(i for i, h in enumerate(header_row) if h and "id nómina" in str(h).lower())
                                idx_fecha_carga = next(i for i, h in enumerate(header_row) if h and "fecha carga" in str(h).lower())
                                idx_fecha_pago = next(i for i, h in enumerate(header_row) if h and "fecha pago" in str(h).lower())
                                idx_estado = next((i for i, h in enumerate(header_row) if h and "estado" in str(h).lower()), None)

                                id_nomina = str(data_row[idx_id]).strip() if idx_id < len(data_row) else None
                                fecha_carga_str = str(data_row[idx_fecha_carga]).strip() if idx_fecha_carga < len(data_row) else None
                                fecha_pago_str = str(data_row[idx_fecha_pago]).strip() if idx_fecha_pago < len(data_row) else None
                                estado = str(data_row[idx_estado]).strip() if idx_estado is not None and idx_estado < len(data_row) else None

                                logger.info(f"ID nómina: {id_nomina}")
                                logger.info(f"Fecha carga (str): {fecha_carga_str}")
                                logger.info(f"Fecha pago (str): {fecha_pago_str}")
                                logger.info(f"Estado: {estado}")

                                # Parsear fechas
                                if fecha_carga_str and fecha_carga_str != "None":
                                    try:
                                        fecha_carga = datetime.strptime(fecha_carga_str, "%d/%m/%Y").date()
                                        logger.info(f"Fecha carga parseada: {fecha_carga}")
                                    except Exception as e:
                                        logger.warning(f"No se pudo parsear Fecha carga '{fecha_carga_str}': {e}")

                                if fecha_pago_str and fecha_pago_str != "None":
                                    try:
                                        fecha_pago = datetime.strptime(fecha_pago_str, "%d/%m/%Y").date()
                                        logger.info(f"Fecha pago parseada: {fecha_pago}")
                                    except Exception as e:
                                        logger.warning(f"No se pudo parsear Fecha pago '{fecha_pago_str}': {e}")

                                if id_nomina:
                                    return {
                                        'id_nomina': id_nomina,
                                        'fecha_carga': fecha_carga,
                                        'fecha_pago': fecha_pago,
                                        'estado': estado
                                    }

                            except (StopIteration, IndexError) as e:
                                logger.warning(f"No se pudieron encontrar todas las columnas: {e}")
                                continue

            raise ValueError("No se encontró ID nómina en las tablas del PDF")

    except Exception as e:
        logger.error(f"Error extrayendo metadatos de nómina: {e}")
        raise


def extraer_ruts_nomina(pdf_path):
    """
    Parsea un PDF de nómina (Estado de Firmas) y extrae los RUTs de beneficiarios.

    Busca patrones de RUT en TODAS las páginas del "Detalle de nómina",
    excluyendo el RUT cliente que aparece en el resumen.
    Devuelve una lista de RUTs únicos de beneficiarios.
    """
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        logger.error(f"PDF no existe: {pdf_path}")
        raise FileNotFoundError(f"PDF no existe: {pdf_path}")

    logger.info(f"Parseando PDF de nómina: {pdf_path}")

    ruts = set()
    rut_cliente = None

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            logger.info(f"PDF abierto: {len(pdf.pages)} páginas")

            # Primera pasada: extraer RUT cliente de la primera página
            if len(pdf.pages) > 0:
                text_pag1 = pdf.pages[0].extract_text()
                if text_pag1 and "RUT cliente" in text_pag1:
                    rut_pattern = r"\d{1,2}\.\d{3}\.\d{3}-[\dkK]|\d{7,8}-[\dkK]|\d{8}[\dkK]"
                    matches = re.findall(rut_pattern, text_pag1)
                    if matches:
                        # El primer RUT después de "RUT cliente" es el cliente
                        rut_cliente = matches[0]
                        logger.info(f"RUT cliente identificado: {rut_cliente}")

            # Segunda pasada: encontrar dónde comienza "Detalle de nómina" y procesar TODAS las páginas desde ahí
            detalle_inicio = -1
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text and "Detalle de nómina" in text:
                    detalle_inicio = page_num
                    logger.info(f"Sección 'Detalle de nómina' encontrada en página {page_num}")
                    break

            # Procesar TODAS las páginas desde "Detalle de nómina" hasta el final
            if detalle_inicio > 0:
                for page_num in range(detalle_inicio, len(pdf.pages) + 1):
                    page = pdf.pages[page_num - 1]
                    text = page.extract_text()

                    if not text:
                        logger.warning(f"No se extrajo texto de página {page_num}")
                        continue

                    logger.info(f"Procesando página {page_num} (detalle de nómina)")

                    # Buscar patrones de RUT
                    rut_pattern = r"\d{1,2}\.\d{3}\.\d{3}-[\dkK]|\d{7,8}-[\dkK]|\d{8}[\dkK]"
                    matches = re.findall(rut_pattern, text)

                    logger.info(f"Encontrados {len(matches)} RUTs en página {page_num}")

                    for rut_str in matches:
                        rut_clean = rut_str.strip()

                        # Excluir el RUT cliente
                        if rut_clean and rut_clean != rut_cliente:
                            ruts.add(rut_clean)
                            logger.debug(f"RUT de beneficiario extraído: {rut_clean}")
            else:
                logger.warning("No se encontró 'Detalle de nómina' en el PDF")

    except Exception as e:
        logger.error(f"Error parseando PDF: {e}")
        raise

    ruts_list = sorted(list(ruts))
    logger.info(f"Se extrajeron {len(ruts_list)} RUTs únicos de beneficiarios: {ruts_list}")

    if not ruts_list:
        raise ValueError("No se extrajeron RUTs de beneficiarios del PDF de nómina")

    return ruts_list


def normalizar_rut(rut):
    """
    Normaliza un RUT: XX.XXX.XXX-X → XXXXXXXX-X
    """
    # Remover puntos
    rut_sin_puntos = rut.replace(".", "")
    return rut_sin_puntos


def extraer_ruts_nomina_excel(excel_path):
    """
    Extrae RUTs de beneficiarios de Excel de nómina (formato CartolaEnLinea).

    Maneja dos formatos:
    - .xlsx moderno (openpyxl)
    - .xls antiguo OLE (xlrd) — el banco descarga como .xlsx pero es realmente .xls antiguo

    Estructura:
    - Hoja: "CartolaEnLinea"
    - Tabla detalle: encabezados en fila 20, datos desde fila 21
    - Columna RUT: "Rut Beneficiario"

    Args:
        excel_path: Ruta al archivo

    Returns:
        Lista de RUTs únicos de beneficiarios
    """
    import openpyxl
    import xlrd

    excel_path = Path(excel_path)

    if not excel_path.exists():
        logger.error(f"Excel no existe: {excel_path}")
        raise FileNotFoundError(f"Excel no existe: {excel_path}")

    logger.info(f"Extrayendo RUTs de Excel: {excel_path}")

    ruts = set()
    rut_pattern = r"\d{1,2}\.\d{3}\.\d{3}-[\dkK]|\d{7,8}-[\dkK]|\d{8}[\dkK]"

    # Intentar primero con openpyxl (.xlsx moderno)
    try:
        logger.info("Intentando leer con openpyxl (.xlsx moderno)...")
        wb = openpyxl.load_workbook(str(excel_path))

        hoja_nombre = None
        for sheet_name in wb.sheetnames:
            if "cartola" in sheet_name.lower() or "nomina" in sheet_name.lower():
                hoja_nombre = sheet_name
                break

        if not hoja_nombre:
            hoja_nombre = wb.sheetnames[0]
            logger.warning(f"Hoja 'CartolaEnLinea' no encontrada, usando: {hoja_nombre}")

        ws = wb[hoja_nombre]
        logger.info(f"Usando hoja con openpyxl: {hoja_nombre}")

        # Leer encabezados de fila 20
        header_row_num = 20
        header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]
        logger.info(f"Encabezados (fila {header_row_num}): {header_row}")

        # Buscar columna "Rut Beneficiario"
        rut_column = None
        for idx, header in enumerate(header_row):
            if header and "rut" in str(header).lower() and "beneficiario" in str(header).lower():
                rut_column = idx
                logger.info(f"Columna 'Rut Beneficiario' en índice {idx}")
                break

        if rut_column is None:
            raise ValueError("No se encontró columna 'Rut Beneficiario'")

        # Extraer RUTs desde fila 21
        data_start_row = 21
        for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not row or not row[rut_column]:
                continue

            cell_str = str(row[rut_column]).strip()
            if cell_str and cell_str.lower() != "rut beneficiario":
                match = re.search(rut_pattern, cell_str)
                if match:
                    rut_encontrado = match.group()
                    ruts.add(rut_encontrado)
                    logger.info(f"Fila {row_num}: {rut_encontrado}")

        wb.close()

    except Exception as e:
        if "File contains no valid workbook part" in str(e):
            logger.warning(f"openpyxl falló: archivo es .xls antiguo OLE. Intentando con xlrd...")

            # Fallback: leer con xlrd (.xls antiguo)
            try:
                wb = xlrd.open_workbook(str(excel_path))

                # Buscar hoja "CartolaEnLinea"
                hoja_nombre = None
                for sheet_name in wb.sheet_names():
                    if "cartola" in sheet_name.lower() or "nomina" in sheet_name.lower():
                        hoja_nombre = sheet_name
                        break

                if not hoja_nombre:
                    hoja_nombre = wb.sheet_names()[0]
                    logger.warning(f"Hoja 'CartolaEnLinea' no encontrada, usando: {hoja_nombre}")

                ws = wb.sheet_by_name(hoja_nombre)
                logger.info(f"Usando hoja con xlrd: {hoja_nombre}")

                # Leer encabezados de fila 20 (0-indexed = 19)
                header_row_idx = 19
                header_row = ws.row_values(header_row_idx)
                logger.info(f"Encabezados (fila {header_row_idx + 1}): {header_row}")

                # Buscar columna "Rut Beneficiario"
                rut_column = None
                for idx, header in enumerate(header_row):
                    if header and "rut" in str(header).lower() and "beneficiario" in str(header).lower():
                        rut_column = idx
                        logger.info(f"Columna 'Rut Beneficiario' en índice {idx}")
                        break

                if rut_column is None:
                    raise ValueError("No se encontró columna 'Rut Beneficiario'")

                # Extraer RUTs desde fila 21 (0-indexed = 20)
                data_start_row = 20
                for row_idx in range(data_start_row, ws.nrows):
                    cell_value = ws.cell_value(row_idx, rut_column)
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if cell_str and cell_str.lower() != "rut beneficiario":
                            match = re.search(rut_pattern, cell_str)
                            if match:
                                rut_encontrado = match.group()
                                ruts.add(rut_encontrado)
                                logger.info(f"Fila {row_idx + 1}: {rut_encontrado}")

            except Exception as e2:
                logger.error(f"Error con xlrd: {e2}", exc_info=True)
                raise
        else:
            logger.error(f"Error inesperado: {e}", exc_info=True)
            raise

    ruts_list = sorted(list(ruts))
    logger.info(f"Se extrajeron {len(ruts_list)} RUTs únicos: {ruts_list}")

    if not ruts_list:
        raise ValueError("No se extrajeron RUTs del Excel de nómina")

    return ruts_list
